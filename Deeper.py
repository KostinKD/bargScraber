import requests
import openpyxl
from bs4 import BeautifulSoup
import time

# Make a GET request to fetch the raw HTML content


ursl = [
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%BD%D1%8B%D0%B9-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81-Time-Park/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B5%D0%B2%D0%B5%D1%80%D0%BD%D1%8B%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8-%D0%B2-%D0%94%D0%BE%D0%BB%D0%B3%D0%BE%D0%BF%D1%80%D1%83%D0%B4%D0%BD%D0%BE%D0%BC/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9A%D0%BB%D1%83%D0%B1-SPACE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/C%D0%B0%D1%83%D0%BD%D0%B0-%D0%9F%D0%B5%D1%80%D0%B2%D0%BE%D0%BC%D0%B0%D0%B9%D1%81%D0%BA%D0%B0%D1%8F-85/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/SPA-%D1%81%D0%B0%D1%83%D0%BD%D0%B0-%D0%92%D0%BE%D0%B7%D1%80%D0%BE%D0%B6%D0%B4%D0%B5%D0%BD%D0%B8%D0%B5/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/VIP-%D0%91%D0%B0%D0%BD%D1%8F-%D0%BD%D0%B0-%D0%B4%D1%80%D0%BE%D0%B2%D0%B0%D1%85-%D0%BD%D0%B0-%D0%9D%D0%BE%D0%B2%D0%BE%D0%BC-%D0%90%D1%80%D0%B1%D0%B0%D1%82%D0%B5-%D0%9C%D0%98%D0%A0%D0%AA-%D0%91%D0%90%D0%9D%D0%98/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/VIP-club-Premium-GALLERY/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%90%D0%BB%D0%B5%D0%BA%D1%81%D0%B0%D0%BD%D0%B4%D1%80%D0%BE%D0%B2%D1%81%D0%BA%D0%B0%D1%8F-%D0%91%D0%B0%D0%BD%D1%8F/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%90%D0%BB%D1%8C%D0%BF%D0%B8%D0%B9%D1%81%D0%BA%D0%B0%D1%8F-%D0%B4%D0%B5%D1%80%D0%B5%D0%B2%D0%BD%D1%8F/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%B8-%D0%BD%D0%B0-%D0%90%D1%81%D1%82%D1%80%D0%B0%D1%85%D0%B0%D0%BD%D1%81%D0%BA%D0%BE%D0%BC-%D0%BF%D0%B5%D1%80%D0%B5%D1%83%D0%BB%D0%BA%D0%B5-%D0%B2-%D0%B3%D0%BE%D1%80%D0%BE%D0%B4%D0%B5-%D0%9C%D0%BE%D1%81%D0%BA%D0%B2%D0%B5/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%B8-%D1%83-%D1%81%D1%82%D0%B0%D0%BD%D1%86%D0%B8%D0%B8-%D0%BC%D0%B5%D1%82%D1%80%D0%BE%D0%BF%D0%BE%D0%BB%D0%B8%D1%82%D0%B5%D0%BD%D0%B0-%D0%91%D0%B0%D0%B1%D1%83%D1%88%D0%BA%D0%B8%D0%BD%D1%81%D0%BA%D0%B0%D1%8F/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D1%8F-%D0%A1%D0%BE%D0%BA%D0%BE%D0%BB%D0%B8%D0%BD%D0%B0%D1%8F-%D0%B3%D0%BE%D1%80%D0%B0/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%B8-%D0%BD%D0%B0-%D0%B2%D0%BE%D0%B4%D0%B5/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%B8-%D0%BF%D0%BE%D1%81%D1%91%D0%BB%D0%BA%D0%B0-%D0%97%D0%B0%D0%B2%D0%B5%D1%82%D1%8B-%D0%98%D0%BB%D1%8C%D0%B8%D1%87%D0%B0/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%BD%D0%BE-%D0%BE%D0%B7%D0%B4%D0%BE%D1%80%D0%BE%D0%B2%D0%B8%D1%82%D0%B5%D0%BB%D1%8C%D0%BD%D1%8B%D0%B9-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81-%D0%BD%D0%B0-%D0%94%D1%83%D0%B1%D0%BD%D0%B8%D0%BD%D1%81%D0%BA%D0%BE%D0%B9/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%BD%D1%8B%D0%B9-%D0%B4%D0%B2%D0%BE%D1%80-%D0%BD%D0%B0-%D0%9B%D0%BE%D0%B1%D0%BD%D0%B5%D0%BD%D1%81%D0%BA%D0%BE%D0%B91/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%BD%D1%8B%D0%B9-%D0%B4%D0%B2%D0%BE%D1%80%D0%B8%D0%BA--%D0%A0%D1%83%D1%87%D0%B5%D1%91%D0%BA/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%94%D0%B5%D0%B2%D1%8F%D1%82%D1%8B%D0%B9-%D0%B2%D0%B0%D0%BB/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%BD%D1%8B%D0%B9-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81-%D0%90%D0%BD%D0%B0%D0%BD%D0%B0%D1%81/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%BD%D1%8B%D0%B9-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81-%D0%93%D0%B5%D0%BD%D0%B5%D1%80%D0%B0%D0%BB%D1%8C%D1%81%D0%BA%D0%B8%D0%B9/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%BD%D1%8B%D0%B9-%D0%BA%D0%BB%D1%83%D0%B1-%D0%9D%D0%B5%D0%BF%D1%82%D1%83%D0%BD/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%BD%D1%8B%D0%B9-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81-%D0%A0%D0%B5%D0%B7%D0%B8%D0%B4%D0%B5%D0%BD%D1%86%D0%B8%D1%8F-%D0%9A%D1%83%D0%BD%D1%86%D0%B5%D0%B2%D0%BE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A8%D0%B0%D0%BB%D0%B5-%D0%BE%D1%82%D0%B5%D0%BB%D1%8C-%D0%A2%D0%B0%D0%B5%D0%B6%D0%BD%D1%8B%D0%B5-%D0%B4%D0%B0%D1%87%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/-%D0%91%D0%B0%D0%BD%D0%BD%D1%8B%D0%B9-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81-%D0%9B%D0%B5%D0%B3%D0%BA%D0%B8%D0%B9-%D0%BF%D0%B0%D1%80/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D1%8F-%D0%90%D0%BA%D0%B2%D0%B0%D0%BB%D0%B8%D0%BD%D0%B0/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D1%8F--2-%D0%96%D0%B0%D1%80-%D0%9F%D1%82%D0%B8%D1%86%D0%B0/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D1%8F-%D0%BD%D0%B0-%D0%B4%D1%80%D0%BE%D0%B2%D0%B0%D1%85-%D0%B2-%D0%96%D0%B5%D0%BB%D0%B5%D0%B7%D0%BD%D0%BE%D0%B4%D0%BE%D1%80%D0%BE%D0%B6%D0%BD%D0%BE%D0%BC/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D1%8F-%D0%BD%D0%B0-%D0%B4%D1%80%D0%BE%D0%B2%D0%B0%D1%85-%D0%B2-%D0%9E%D1%81%D1%82%D0%B0%D1%84%D1%8C%D0%B5%D0%B2%D0%BE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D1%8F-%D0%BD%D0%B0-%D0%9C%D0%BE%D1%80%D0%BE%D0%B7%D0%B5/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D1%8F-%D0%BD%D0%B0-%D0%A1%D0%BE%D0%B2%D1%85%D0%BE%D0%B7%D0%BD%D0%BE%D0%B9-%D1%83%D0%BB-8-/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D1%8F-%D0%A8%D0%B0%D1%85%D0%BE%D0%B2%D1%81%D0%BA%D0%B0%D1%8F/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D1%81%D0%BA%D0%B5%D1%82-%D0%91%D0%B0%D1%80/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B5%D0%BB%D1%8B%D0%B5-%D0%A1%D1%82%D0%BE%D0%BB%D0%B1%D1%8B-%D0%B2-%D0%94%D0%BE%D0%BC%D0%BE%D0%B4%D0%B5%D0%B4%D0%BE%D0%B2%D0%BE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D1%8B%D0%BA%D0%BE%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%92%D0%B8%D0%B4%D0%BD%D0%BE%D0%B2%D1%81%D0%BA%D0%B0%D1%8F-%D0%B3%D0%BE%D1%80%D0%BE%D0%B4%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%92%D0%BD%D1%83%D0%BA%D0%BE%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%92%D0%BE%D1%80%D0%BE%D0%BD%D1%86%D0%BE%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%92%D0%BE%D1%81%D1%82%D0%BE%D1%87%D0%BD%D1%8B%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8-%D0%B2-%D0%BF%D0%BE%D1%81%D0%B5%D0%BB%D0%BA%D0%B5-%D0%92%D0%BE%D1%81%D1%82%D0%BE%D1%87%D0%BD%D0%BE%D0%BC/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%92%D0%BE%D1%81%D1%82%D1%80%D1%8F%D0%BA%D0%BE%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%92%D1%83%D0%B3%D0%B8-%D0%92%D1%83%D0%B3%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%92%D1%8F%D1%82%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%93%D0%BE%D1%80%D0%BE%D0%B4%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F-%D0%B2-%D0%94%D0%B7%D0%B5%D1%80%D0%B6%D0%B8%D0%BD%D1%81%D0%BA%D0%BE%D0%BC/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%93%D0%BE%D1%80%D0%BE%D0%B4%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F-%D0%B2-%D0%9E%D0%B4%D0%B8%D0%BD%D1%86%D0%BE%D0%B2%D0%BE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%94%D0%B5%D0%B4%D0%BE%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%94%D0%BE%D0%B7%D0%B0%D0%BF%D1%80%D0%B0%D0%B2%D0%BA%D0%B0/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%94%D0%BE%D0%BB%D0%B3%D0%BE%D0%BF%D1%80%D1%83%D0%B4%D0%BD%D0%B5%D0%BD%D1%81%D0%BA%D0%B0%D1%8F-%D0%B3%D0%BE%D1%80%D0%BE%D0%B4%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%96%D1%83%D0%BA%D0%BE%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9A%D0%B0%D0%BB%D0%B8%D1%82%D0%BD%D0%B8%D0%BA%D0%BE%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9A%D0%B0%D0%BF%D0%BE%D1%82%D0%BD%D0%B8%D0%BD%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9A%D0%B5%D0%B4%D1%80%D0%BE%D0%B2%D1%8B%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9A%D0%BE%D1%81%D0%B8%D0%BD%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9A%D1%80%D0%B0%D1%81%D0%BD%D0%BE%D0%BF%D1%80%D0%B5%D1%81%D0%BD%D0%B5%D0%BD%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9A%D1%83%D0%BD%D1%86%D0%B5%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8-/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9B%D0%B5%D1%84%D0%BE%D1%80%D1%82%D0%BE%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9B%D1%83%D0%B3%D0%BE%D0%B2%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9B%D1%8B%D1%82%D0%BA%D0%B0%D1%80%D0%B8%D0%BD%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9B%D1%8C%D0%B2%D0%BE%D0%B2%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9B%D1%8E%D0%B1%D0%BB%D0%B8%D0%BD%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9C%D0%B0%D1%80%D1%8C%D0%B8%D0%BD%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8-%D0%BD%D0%B0-%D1%83%D0%BB%D0%B8%D1%86%D0%B5-%D0%9D%D0%B8%D0%B6%D0%BD%D0%B8%D0%B5-%D0%9F%D0%BE%D0%BB%D1%8F/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9C%D1%8B%D1%82%D0%B8%D1%89%D0%B8%D0%BD%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9D%D0%B0%D0%B7%D0%B0%D1%80%D1%8C%D0%B5%D0%B2%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9D%D0%B0%D1%85%D0%B0%D0%B1%D0%B8%D0%BD%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9E%D0%B7%D0%B4%D0%BE%D1%80%D0%BE%D0%B2%D0%B8%D1%82%D0%B5%D0%BB%D1%8C%D0%BD%D0%BE-%D0%B1%D0%B0%D0%BD%D0%BD%D1%8B%D0%B9-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81-%D0%9D%D0%B5%D0%BA%D1%80%D0%B0%D1%81%D0%BE%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%91%D0%B0%D0%BD%D0%B8-%D0%B2-%D0%9D%D0%B5%D0%BA%D1%80%D0%B0%D1%81%D0%BE%D0%B2%D0%BA%D0%B5/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9D%D0%B8%D0%BA%D0%BE%D0%BB%D0%B0%D0%B5%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9D%D0%BE%D0%B2%D0%BE%D0%BD%D0%B8%D0%BA%D0%BE%D0%BB%D1%8C%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9E%D0%B1%D1%89%D0%B5%D1%81%D1%82%D0%B2%D0%B5%D0%BD%D0%BD%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F--1-%D0%B2-%D0%9C%D0%B0%D0%BB%D0%B0%D1%85%D0%BE%D0%B2%D0%BA%D0%B5/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9E%D0%B1%D1%89%D0%B5%D1%81%D1%82%D0%B2%D0%B5%D0%BD%D0%BD%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F-%D0%B2-%D0%9B%D1%8B%D1%82%D0%BA%D0%B0%D1%80%D0%B8%D0%BD%D0%BE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9E%D0%B1%D1%89%D0%B5%D1%81%D1%82%D0%B2%D0%B5%D0%BD%D0%BD%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F-%D0%B2-%D0%9F%D0%BE%D0%B4%D0%BE%D0%BB%D1%8C%D1%81%D0%BA%D0%B5-%D0%91%D0%B0%D0%BD%D1%8C%D0%BA%D0%B0/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9E%D0%B1%D1%89%D0%B5%D1%81%D1%82%D0%B2%D0%B5%D0%BD%D0%BD%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F-%D0%BD%D0%B0-%D0%A1%D0%BE%D0%B2%D1%85%D0%BE%D0%B7%D0%BD%D0%BE%D0%B9-%D1%83%D0%BB-55/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9E%D0%B1%D1%89%D0%B5%D1%81%D1%82%D0%B2%D0%B5%D0%BD%D0%BD%D0%B0%D1%8F-%D1%80%D1%83%D1%81%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F-%D0%B2-%D0%9F%D0%BE%D0%B4%D0%BE%D0%BB%D1%8C%D1%81%D0%BA%D0%B5/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9E%D0%B1%D1%89%D0%B5%D1%81%D1%82%D0%B2%D0%B5%D0%BD%D0%BD%D1%8B%D0%B5-%D0%9A%D0%BE%D1%80%D0%BE%D0%BB%D1%91%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D1%8F-%D1%81%D0%B0%D1%83%D0%BD%D0%B0-%D0%BD%D0%B0-%D0%9E%D1%87%D0%B0%D0%BA%D0%BE%D0%B2%D1%81%D0%BA%D0%BE%D0%B9/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9F%D0%B5%D1%80%D0%BB%D0%BE%D0%B2%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F-%D0%B2-%D0%9C%D1%8B%D1%82%D0%B8%D1%89%D0%B0%D1%85/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9F%D0%BE%D0%B4%D0%BE%D0%BB%D1%8C%D1%81%D0%BA%D0%B8%D0%B5-%D0%B3%D0%BE%D1%80%D0%BE%D0%B4%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9F%D0%BE%D0%BA%D1%80%D0%BE%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%BD%D1%8B%D0%B9-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81-%D0%A0%D0%B5%D1%83%D1%82%D0%BE%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8-%D0%B8-%D1%81%D0%B0%D1%83%D0%BD%D1%8B/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A0%D0%B6%D0%B5%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8--%D0%BA%D0%B0%D1%87%D0%B5%D1%81%D1%82%D0%B2%D0%BE-%D0%BF%D1%80%D0%BE%D0%B2%D0%B5%D1%80%D0%B5%D0%BD%D0%BD%D0%BE%D0%B5-%D0%B2%D1%80%D0%B5%D0%BC%D0%B5%D0%BD%D0%B5%D0%BC/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A0%D1%83%D1%81%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F-%D0%B2-%D0%94%D1%83%D0%B1%D1%80%D0%BE%D0%B2%D0%B8%D1%86%D0%B0%D1%85/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A0%D1%83%D1%81%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F-%D0%B2-%D0%9F%D0%BE%D0%B4%D0%BE%D0%BB%D1%8C%D1%81%D0%BA%D0%B5/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A0%D1%83%D1%81%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F-%D0%BD%D0%B0-%D0%B4%D1%80%D0%BE%D0%B2%D0%B0%D1%85-555/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A0%D1%83%D1%81%D1%81%D0%BA%D0%B0%D1%8F-%D0%B1%D0%B0%D0%BD%D1%8F-%D0%BD%D0%B0-%D0%B4%D1%80%D0%BE%D0%B2%D0%B0%D1%85-%D0%B2-%D0%90%D0%BA%D1%81%D0%B8%D0%BD%D1%8C%D0%B8%D0%BD%D0%BE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A0%D1%83%D1%81%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8-%D0%B2-%D0%91%D0%B0%D0%BB%D0%B0%D1%88%D0%B8%D1%85%D0%B5/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%91%D0%B0%D0%BD%D0%BD%D1%8B%D0%B9-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81-%D0%A1%D0%B0%D0%BD%D0%B4%D1%83%D0%BD%D0%BE%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%93%D0%9E%D0%90/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%9A%D0%BE%D0%BB%D0%B8%D0%B1%D1%80%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%9D%D0%B0%D0%B4%D0%B5%D0%B6%D0%B4%D0%B0/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%9D%D0%B8%D1%80%D0%B2%D0%B0%D0%BD%D0%B0-%D0%97%D0%B0%D0%BB-%D0%9A%D0%BE%D1%80%D0%B0%D0%B1%D0%BB%D1%8C/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%9D%D0%BE%D0%B2%D1%8B%D0%B9-%D0%BC%D0%B8%D1%80/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%9E%D1%81%D1%82%D1%80%D0%BE%D0%B2/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%9F%D0%B0%D0%BB%D1%8C%D0%BC%D0%B8%D1%80%D0%B0/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%96%D0%B5%D0%BC%D1%87%D1%83%D0%B6%D0%B8%D0%BD%D0%B0-%D0%9D%D0%B8%D0%BB%D0%B0-/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-BonApart/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%9F%D0%BE-%D0%A6%D0%B0%D1%80%D1%81%D0%BA%D0%B8-%D0%9F%D0%B5%D1%80%D0%BE%D0%B2%D0%BE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%91%D1%83%D1%82%D0%BE%D0%B2%D0%BE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%B8-%D1%85%D0%B0%D0%BC%D0%BC%D0%B0%D0%BC-%D0%B2-%D0%9C%D0%B0%D0%BA%D1%81%D0%B8%D0%BC%D0%B0-%D0%A1%D0%BB%D0%B0%D0%B2%D0%B8%D1%8F-%D0%9E%D1%82%D0%B5%D0%BB%D1%8C/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%BD%D0%B0-%D0%9B%D0%B8%D0%B4%D0%B5%D1%80-%D0%BD%D0%B0-%D0%9F%D0%B5%D1%80%D0%B2%D0%BE%D0%BC%D0%B0%D0%B9%D1%81%D0%BA%D0%BE%D0%B9/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%BD%D0%B0-%D0%91%D0%B0%D0%B9%D0%BA%D0%B0%D0%BB%D1%8C%D1%81%D0%BA%D0%BE%D0%B9/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%BD%D0%B0-%D0%9B%D0%B8%D1%85%D0%BE%D0%B1%D0%BE%D1%80%D0%B0%D1%85/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%BD%D0%B0-%D0%A0%D1%8F%D0%B7%D0%B0%D0%BD%D1%81%D0%BA%D0%BE%D0%BC-%D0%BF%D1%80%D0%BE%D1%81%D0%BF%D0%B5%D0%BA%D1%82%D0%B5/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%BD%D0%B0-%D0%A4%D1%80%D1%83%D0%BD%D0%B7%D0%B5%D0%BD%D1%81%D0%BA%D0%BE%D0%B9/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%BD%D0%B0-%D0%A8%D0%B0%D0%B1%D0%BE%D0%BB%D0%BE%D0%B2%D0%BA%D0%B5/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%9E%D1%81%D1%82%D0%B0%D0%BD%D0%BA%D0%B8%D0%BD%D0%BE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%A0%D0%B5%D1%87%D0%BD%D0%BE%D0%B9-%D0%B2%D0%BE%D0%BA%D0%B7%D0%B0%D0%BB/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%A4%D0%B5%D0%BD%D0%B8%D0%BA%D1%81/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%BA%D0%B0%D1%84%D0%B5-%D0%A1%D1%83%D0%BC%D0%B0%D1%85/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%BA%D0%BB%D1%83%D0%B1-%D0%9C%D0%B0%D1%80%D1%80%D0%B0%D0%BA%D0%B5%D1%88/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B5%D0%BB%D0%B5%D0%B7%D0%BD%D0%B5%D0%B2%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%BB%D0%B0%D0%B2%D1%8F%D0%BD%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8-%D0%B2-%D0%A9%D1%91%D0%BB%D0%BA%D0%BE%D0%B2%D0%BE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%BB%D0%BE%D0%B1%D0%BE%D0%B4%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%BE%D0%B2%D1%85%D0%BE%D0%B7%D0%BD%D1%8B%D0%B9-%D0%B1%D0%B0%D0%BD%D0%BD%D1%8B%D0%B9-%D0%BA%D0%BE%D0%BC%D0%BF%D0%BB%D0%B5%D0%BA%D1%81/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A1%D0%B0%D1%83%D0%BD%D0%B0-%D0%9B%D1%8E%D0%BA%D1%81-%D0%B2-%D0%A7%D0%B5%D1%80%D1%82%D0%B0%D0%BD%D0%BE%D0%B2%D0%BE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A2%D1%83%D1%80%D0%B5%D1%86%D0%BA%D0%B8%D0%B9-%D1%85%D0%B0%D0%BC%D0%BC%D0%B0%D0%BC-%D0%B2-%D0%9C%D0%B0%D0%BA%D1%81%D0%B8%D0%BC%D0%B0-%D0%98%D1%80%D0%B1%D0%B8%D1%81-%D0%9E%D1%82%D0%B5%D0%BB%D1%8C/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A3%D1%81%D0%B0%D0%B4%D1%8C%D0%B1%D0%B0-%D0%B1%D0%B0%D0%BD%D0%BD%D0%B0%D1%8F-%D0%BD%D0%B0-%D0%92%D0%B8%D1%88%D0%BD%D0%B5%D0%B2%D0%BE%D0%B9/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A3%D1%81%D0%B0%D0%B4%D1%8C%D0%B1%D0%B0-%D0%B1%D0%B0%D0%BD%D0%BD%D0%B0%D1%8F-%D0%BD%D0%B0-%D0%92%D0%BE%D0%B9%D0%BA%D0%BE%D0%B2%D1%81%D0%BA%D0%BE%D0%B9/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A3%D1%81%D0%B0%D0%B4%D1%8C%D0%B1%D0%B0-%D0%91%D0%B0%D0%BD%D0%BD%D0%B0%D1%8F/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A3%D1%81%D0%B0%D0%B4%D1%8C%D0%B1%D0%B0-%D0%91%D0%B0%D0%BD%D0%BD%D0%B0%D1%8F-%D0%BD%D0%B0-%D0%9E%D0%BA%D1%82%D1%8F%D0%B1%D1%80%D1%8C%D1%81%D0%BA%D0%BE%D0%BC-%D0%BF%D0%BE%D0%BB%D0%B5/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A3%D1%81%D0%B0%D0%B4%D1%8C%D0%B1%D0%B0-%D0%91%D0%B0%D0%BD%D0%BD%D0%B0%D1%8F-%D0%BD%D0%B0-%D0%9F%D0%B0%D0%BD%D1%84%D0%B8%D0%BB%D0%BE%D0%B2%D0%B0-/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A3%D1%81%D0%B0%D0%B4%D1%8C%D0%B1%D0%B0-%D0%B2-%D0%AE%D1%80%D0%BB%D0%BE%D0%B2%D0%BE/',
'https://бани.рф/%D0%B1%D0%B0%D0%BD%D1%8F/%D0%A6%D0%B0%D1%80%D0%B8%D1%86%D1%8B%D0%BD%D1%81%D0%BA%D0%B8%D0%B5-%D0%B1%D0%B0%D0%BD%D0%B8-%D0%B2-%D0%A2%D0%A6-%D0%A6%D0%B0%D1%80%D0%B8%D1%86%D1%8B%D0%BD%D0%BE/'
        ]

# Москва
# Санкт-Петербург
# Волгоград
# Екатеринбург
# Казань
# Нижний Новгород
# Новосибирск
# Омск
# Ростов на Дону
# Самара
# Уфа
# Челябинск

with open('./Omsk/URLs.txt','r', encoding="utf-8") as f:
    lines = f.read().split(',')

ursl = [line.strip() for line in lines]
# print(ursl)


with open('address.txt','w') as file:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for idx,url in enumerate(ursl):
        print(idx+1,url)
        # print(ursl)

        page = requests.get(url)
    # Parse the html content
        soup = BeautifulSoup(page.content, "html.parser")

        # addr_list = soup.find_all('div', attrs={'class':'addrsmall3'})
        name_bath = soup.find('div', {'id':'contnt'}).h1.text
        try:
            addr_list = soup.find('div', class_='addrsmall3').text
        except Exception as e:
            print(e)
            continue
        descript_list = soup.find('div', class_='pad1det').text
        price_list = soup.find('div', class_='addrsmall5').text
        phone_list = soup.find('div', class_='phone2').text
        service_list = soup.find('div', class_='formserv').text
        service_list2 = soup.find_all(class_='formserv')
        service_list3 = soup.find("span", class_= "servgrp", string="Кухня:")
        # service_list4 = soup.find("span", class_= "servgrp").text
        service_a = soup.find('div', class_='formserv').a


        # nextele = service_a.find_next_siblings("a")
        # for next in nextele:
        #     print(next.text)

        test_all = soup.find('div', class_='formserv').find_all()



        Type_Bath = []
        Type_Kitchen = []
        Type_Service = []
        Type_Services = []
        start = False
        for element in reversed(test_all):
            print(element)
            start = True
            ########################### SERVICE СЕРВИС (WORK) -> addd reversed to test_all
            # if element.text == 'Сервис:':
            #     start = False
            #     if len(Type_Services) > 1:
            #         Type_Services.reverse()
            #     Type_Services = ', '.join(Type_Services)
            #     print(Type_Services)
            #     worksheet.cell(row=idx + 1, column=9).value = Type_Services
            #     break
            # if start:
            #     Type_Services.append(element.text)
            #     print(Type_Services)

            # Type_Bath = []
            # # Scrabing TYPE OF BATHS
            # if element.text == 'Кухня:':
            #     print("Нашли span с текстом 'Кухня', останавливаемся")
            #     if len(Type_Bath) > 1:
            #         Type_Bath.pop(0)
            #         Type_Bath.pop(-1)
            #     Type_Bath = ', '.join(Type_Bath)
            #     worksheet.cell(row=idx + 1, column=9).value = Type_Services
            #     break
            # else:
            #     print(element.text)
            #     Type_Bath.append(element.text)
            #     print(Type_Bath)



            ###########################TYPE OF BATH(WORK)
            if element.text in ['Кухня:', 'Услуги:', 'Сервис:']:
                # print("Нашли span с текстом 'Кухня', останавливаемся")
                if len(Type_Bath) > 1:
                    Type_Bath.pop(0)
                    Type_Bath.pop(-1)
                Type_Bath = ', '.join(Type_Bath)
                # worksheet.cell(row=idx + 1, column=7).value = Type_Bath
                break
            else:
                # print(element.text)
                Type_Bath.append(element.text)
                print(Type_Bath)
            ###############################

            # for element in test_all:
            #     print(element)
            #     # Type_Bath = []
            #     # if element.find('span') != None or element.find('span').text == 'Кухня:':
            #     if element.text == 'Кухня:':
            #         print("Нашли span с текстом 'Кухня', останавливаемся")
            #         Type_Bath.pop(0)
            #         Type_Bath.pop(-1)
            #         if len(Type_Bath) > 1:
            #             Type_Bath.pop(0)
            #             Type_Bath.pop(-1)
            #         Type_Bath = ', '.join(Type_Bath)
            #         worksheet.cell(row=idx + 1, column=7).value = Type_Bath
            #         break
            #     else:
            #         print(element.text)
            #         Type_Bath.append(element.text)
            #         print(Type_Bath)

        ################SCARGIN KITCHEN(WORK)
            # if element.text == 'Кухня:':
            #     print('Scrab kitchen now')
            #     start = True
            # elif element.text == 'Услуги:':
            #     start = False
            #     if len(Type_Kitchen) > 1:
            #         Type_Kitchen.pop(0)
            #         Type_Kitchen.pop(-1)
            #     Type_Kitchen = ', '.join(Type_Kitchen)
            #     print(Type_Kitchen)
            #     # worksheet.cell(row=idx + 1, column=9).value = Type_Kitchen
            #     break
            # if start:
            #     Type_Kitchen.append(element.text)
            #     print(Type_Kitchen)
        #############################


        ################# SCRABER SERVICE(WORK) - УСЛУГИ
            # if element.text == 'Услуги:':
            #     print('Scrab service now')
            #     start = True
            # elif element.text == 'Сервис:':
            #     start = False
            #     if len(Type_Service) > 1:
            #         Type_Service.pop(0)
            #         Type_Service.pop(-1)
            #     Type_Service = ', '.join(Type_Service)
            #     print(Type_Service)
            #     worksheet.cell(row=idx + 1, column=8).value = Type_Service
            #     break
            # if start:
            #     Type_Service.append(element.text)
            #     print(Type_Service)

            # for i in range(len(service_list2)):
            #     if i == len(service_list2) - 1:
            #         result = service_list2[i].next_sibling.strip()
            #         print(result)
            #     else:
            #         result = service_list2[i].next_sibling
            #         while result and result.name == 'br':
            #             result = result.next_sibling
            #         if result:
            #             result = result.strip()
            #     print(result)
            #     worksheet.cell(row=idx + 1, column=6).value = result

        # content = ''
        # kitchen_span = soup.find('span', text='Кухня:')
        # if kitchen_span:
        #     kitchen_div = kitchen_span.find_parent('div')
        #     content = kitchen_div.previous_sibling
        #     while content and content.name != 'Кухня':
        #         if hasattr(content, 'text'):
        #             print(content.text.strip())
        #         content = content.previous_sibling


        details_list = soup.find('div', class_='detailright1').text

        tabs_list = soup.find('div', class_='zalpads')
        many = False
        if tabs_list:
            many = True
        else:
            pass


        worksheet.cell(row=idx+1, column=1).value = name_bath
        worksheet.cell(row=idx+1, column=2).value = url
        worksheet.cell(row=idx+1, column=3).value = addr_list
        worksheet.cell(row=idx+1, column=4).value = descript_list
        try:
            worksheet.cell(row=idx+1, column=5).value = Type_Bath
        except Exception as e:
            print(e)
            continue
        worksheet.cell(row=idx+1, column=6).value = details_list
        # worksheet.cell(row=idx+1, column=7).value = many
        # worksheet.cell(row=idx+1, column=8).value = Type_Bath
        worksheet.cell(row=idx+1, column=10).value = price_list
        worksheet.cell(row=idx+1, column=11).value = phone_list


        # try:
        #     worksheet.cell(row=idx + 1, column=9).value = Type_Kitchen
        # except Exception as e:
        #     print(e)
        #     continue

        time.sleep(0.5)
        workbook.save(filename="output2.xlsx")



    # div_script_list = soup.find('div', class = 'banyabig')

        # print(addr_list)

        file.write(addr_list+ '\n')
    # Get the list of all cities
    #     bath_list = soup.find_all('div', attrs={'class':'banyabig'})

        # for div in bath_list:
        # # bath_price = div.find('div', class_='starn sel').text
        #     bath_stars = len(div.find_all('div', {'class': 'starn sel'}))
        #     print(bath_address)


