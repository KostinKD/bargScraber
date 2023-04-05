import requests
import openpyxl
from bs4 import BeautifulSoup
import time

# Make a GET request to fetch the raw HTML content




url = 'https://xn--90ahkico2a6b9d.xn--80abwt.xn--p1ai/list.html'
page = requests.get(url)
# Parse the html content
soup = BeautifulSoup(page.content, "html.parser")


# div_script_list = soup.find('div', class = 'banyabig')

arr = [ 1 ,2 ]
arr2 = arr.__len__()
# Get the list of all cities
bath_list = soup.find_all('div', attrs={'class':'banyabig'})
workbook = openpyxl.Workbook()
worksheet = workbook.active

for idx,div in enumerate(bath_list):
    bath_name= div.find('div', class_='bname2').text
    # bath_price = div.find('div', class_='starn sel').text
    bath_stars = len(div.find_all('div', {'class': 'starn sel'}))

    worksheet.cell(row=idx + 1, column=1).value = bath_name
    worksheet.cell(row=idx + 1, column=2).value = bath_stars
    workbook.save(filename="stars.xlsx")
    print(bath_name)
    print(bath_stars)