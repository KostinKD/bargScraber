import openpyxl
import time

# Открываем файл Excel
workbook = openpyxl.load_workbook('./NinzniyNovgorod/NizniyNovgorod.xlsx')
# Выбираем лист, на котором расположены данные
worksheet = workbook.active

# Читаем значения ячеек в строке и преобразуем в массив

Type_Bath = {
'финская парная' : 1,
'русская баня': 2,
'русская на дровах': 3,
'турецкая парная(хамам)': 4,
'инфракрасная сауна': 3,
}

Type_Zones = {
'бассейн' : 	1,
'охраняемая парковка' : 	2,
'холодная купель' : 	3,
'обливное ведро' : 	4,
'спутниковое тв' : 	5,
'кондиционер' : 	6,
'бильярд' : 	7,
'кондиционер' : 	8,
'бильярд' : 	9,
'холодная купель' : 	10,
'охраняемая парковка' : 	11,
'прорубь' : 	12,
'спутниковое тв' : 	13,
'большой TV' : 	14,
'обливное ведро' : 	15,
'настольные игры' : 	16,
'джакузи' : 	17,
'Wi-Fi' : 	18,
'караоке' : 	19,
'японский офуро' : 	20,
'загородный отдых' : 	21,
'отель' : 	22,
'бассейн спротивотоком' : 	23,
'камин' : 	24
}


def bath_steam_room():
    row = 1
    for idx, cell in enumerate(worksheet['E'], start=1):
        try:
            values = [x.strip() for x in cell.value.split(',')]
            print('Что в ячейке(массив): ', values)
            print('Чистые значения: ', cell.value)
            for words in values:
                if words in Type_Bath:
                    print('Совпало в словаре')
                    worksheet.cell(row=row, column=16).value = Type_Bath[words]
                    worksheet.cell(row=row, column=15).value = idx
                    print('ID в словаре: ', Type_Bath[words])
                    row += 1

            idx += 1
            print('Индекс: ', idx)
        except Exception as e:
            print(e)
            continue

        time.sleep(0.5)
        workbook.save(filename="DBConvert.xlsx")


def bath_zone():
    row = 1
    for idx, cell in enumerate(worksheet['H'], start=1):
        try:
            values = [x.strip() for x in cell.value.split(',')]
            print('Что в ячейке(массив): ', values)
            print('Чистые значения: ', cell.value)
            for words in values:
                if words in Type_Zones:
                    print('Совпало в словаре')
                    worksheet.cell(row=row, column=18).value = idx
                    worksheet.cell(row=row, column=19).value = Type_Zones[words]
                    print('ID в словаре: ', Type_Zones[words])
                    row += 1

            idx += 1
            print('Индекс: ', idx)
        except Exception as e:
            print(e)
            continue
        time.sleep(0.5)
        workbook.save(filename="DBImport/DBZones.xlsx")




bath_zone()

